from decimal import Decimal, getcontext
import locale
import openpyxl
import re

# Set the desired precision for currency with 2 decimal places
getcontext().prec = 2
locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')

def extract_employee_info(dirpath, filename, heading):
    filepath = dirpath + filename
    col_b_index = openpyxl.utils.column_index_from_string('B')
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    # Find the "Employees" heading and get the row number
    employees_heading = next((cell.row for row in sheet.iter_rows(min_row=1, max_row=698, min_col=1, max_col=col_b_index)
                              for cell in row if cell.value == heading), None)
    if not employees_heading:
        print("The 'Employees' heading was not found in the worksheet.")
        return [], []

    heading_list, filtered_employee_data = [], []
    start_extracting = False

    for row in sheet.iter_rows(min_row=employees_heading, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        current_employee = []

        if any(cell.value == "NetPay" and cell.value != 'Grand Totals' for cell in row):
            start_extracting = False
            break
        if start_extracting:
            for cell in row:
                if cell.coordinate in sheet.merged_cells and sheet[cell.coordinate].value:
                    merged_cell_value = re.sub(r"\n+", ", ", "First Name: " + sheet[cell.coordinate].value)
                    if merged_cell_value != 'First Name: Grand Totals':
                        split_string_list = merged_cell_value.split(",")
                        current_employee.append(split_string_list)
                        filtered_employee_data.append([info.split(":")[1] for info in split_string_list if len(info.split(":")) == 2])
            heading_list = [info.split(":")[0] for info in split_string_list if len(info.split(":")) == 2]
        if any(cell.value for cell in row):
            start_extracting = True

        if not any(cell.value for cell in row):
            break
    return heading_list, filtered_employee_data

def payroll_taxes(dirpath, filename, heading):
    filepath = dirpath + filename
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    # Find the row containing the heading
    heading_row = next((cell.row for row in sheet.iter_rows() for cell in row if cell.value == heading), None)
    if heading_row is None:
        print(f"Heading '{heading}' not found.")
        return [], []

    col_j_index = openpyxl.utils.column_index_from_string('J')
    col_i_index = openpyxl.utils.column_index_from_string('I')
    row_index = heading_row + 1
    __taxes, __all_taxes, __tax_headings = [], [], []

    while True:
        cell_value = sheet.cell(row=row_index, column=col_j_index).value
        cell_tax_title = sheet.cell(row=row_index, column=col_i_index).value

        if cell_value is None:
            return __tax_headings, __all_taxes

        if cell_tax_title == 'CA California state tax':
            __all_taxes.append(__taxes)
            __taxes = []
            row_index += 4

        if cell_value is not None:
            if cell_tax_title not in __tax_headings:
                __tax_headings.append(cell_tax_title)
            __taxes.append(f"${cell_value:,.2f}")
            row_index += 1

def extract_taxes_paid(dirpath, filename, heading):
    filepath = dirpath + filename
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    # Find the row containing the heading
    heading_row = next((cell.row for row in sheet.iter_rows() for cell in row if cell.value == heading), None)
    if heading_row is None:
        print(f"Heading '{heading}' not found.")
        return []

    col_j_index = openpyxl.utils.column_index_from_string('J')
    row_index = heading_row + 10
    __taxes = []

    while True:
        cell_value = sheet.cell(row=row_index, column=col_j_index).value
        if cell_value is None:
            break
        if cell_value != heading:
            taxes = locale.currency(cell_value, grouping=True)
            __taxes.append(taxes)
        row_index += 10

    return __taxes

def combine_list(x, y, z):
    return x + y + [z]

def combine_data_lists(x, y, z):
    return [x[i] + y[i] + [z[i]] for i in range(len(x))]

def write_to_file_in_directory(dirpath, filename, headings, emp_data):
    filename = re.sub("xlsx", "csv", filename)
    filepath = dirpath + filename
    emp_data.insert(0, headings)

    with open(filepath, 'w') as file:
        for line in emp_data:
            file.write(str(line) + '\n')
    print("Data processes and written to ${filepath}") 

if __name__ == '__main__':
    v_path = 'data\\'
    v_file = 'Zenefits_payroll_detail_Ck Date 121523.xlsx'

    l_emp_info_heading, l_emp_info = extract_employee_info(v_path, v_file, heading='Employees')
    l_payroll_taxes_headings, l_payroll_taxes = payroll_taxes(v_path, v_file, heading='Amount')
    l_taxes_paid = extract_taxes_paid(v_path, v_file, heading='Amount')
    v_taxes_paid_heading = 'Taxes Paid'

    l_combined_headings = combine_list(l_emp_info_heading, l_payroll_taxes_headings, v_taxes_paid_heading)
    l_combined_emp_tax = combine_data_lists(l_emp_info, l_payroll_taxes, l_taxes_paid)

    write_to_file_in_directory(v_path, v_file, l_combined_headings, l_combined_emp_tax)
