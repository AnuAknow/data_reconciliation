from decimal import getcontext
import pandas as pd 
import numpy as np
import locale
import openpyxl
import sys
import re
import os


# Set the desired precision (e.g., for typical currency with 2 decimal places)
getcontext().prec = 2

locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')

"""
This Python script accepts user input and checks if it's a non-empty string. 
If the input is not a valid string or is empty, it will ask for input up to 3 
times before raising an error

This script defines a get_filepath_input and a get_filename_input function that:
Prompts the user for input. Checks if the input is a non-empty string.
If the input is invalid, it asks for input again up to 3 more times.
If the input is still invalid after 3 more attempts, it raises a ValueError.
"""
def get_filepath_input(prompt):
    attempts = 0
    max_attempts = 3
    
    while attempts <= max_attempts:
        user_input = input(prompt)
        
        if isinstance(user_input, str) and user_input.strip():
            return user_input
        else:
            attempts += 1
            print(f"Invalid input. Please enter a non-empty string. Attempts remaining: {max_attempts - attempts + 1}")
            
    raise ValueError("Maximum attempts exceeded. Valid input not provided.")

def get_filename_input(prompt):
    attempts = 0
    max_attempts = 3
    
    while attempts <= max_attempts:
        user_input = input(prompt)
        
        if isinstance(user_input, str) and user_input.strip():
            return user_input
        else:
            attempts += 1
            print(f"Invalid input. Please enter a non-empty string. Attempts remaining: {max_attempts - attempts + 1}")
            
    raise ValueError("Maximum attempts exceeded. Valid input not provided.")

def get_sort_row_index(prompt):
    attempts = 0
    max_attempts = 3

    while attempts < max_attempts:
        try:
            input_index = input(prompt)
            input_index = input_index.strip()

            if not input_index.isdigit():
                raise ValueError("You must enter a single numeric value.")

            sort_column_index = int(input_index)
            print(f"Valid sort column index entered: {sort_column_index}")
            return sort_column_index
        
        except ValueError as e:
            attempts += 1
            print(f"Invalid input: {e}")
            if attempts < max_attempts:
                print(f"You have {max_attempts - attempts} attempts left.")
            else:
                print("No more attempts left. Please restart the program.")

def extract_employee_info(file_path, filename, heading, start_row, stop_row):
    """
    This python function extracts data by:
    
    1. locating a heading and returning the column index 
    2. it extracts the values from 10 rows from a 10 row merged cell object 
       beneath the located heading 
    3. After extracting 10 rows it advances 1 row and continues to extracting 
       values from the next 10 row merged cell object until it reach the end_row 
    
    This script defines an extract_data function that takes the following parameters:

    Args:
        file_path: The path to your Excel file.
        heading: The heading to identify the column to extract data from.
        start_row: The row where the table heading is located.
        num_rows: The number of rows to extract below each heading.
        skip_rows: The number of rows to skip after extracting.
        end_row: The row to stop extraction.

    This function extracts the specified number of rows below each heading, skips the given
    number of rows, and continues this pattern until it reaches the end row. You can adjust 
    the parameters to fit your specific needs.
    
    """
    # Get the current working directory
    current_directory = os.getcwd()
    print(f"Current directory: {current_directory}")
    
    try:
        # Get the current working directory
        directory_list = current_directory.split("\\")
        if file_path not in directory_list:
            print(f"No {file_path} is in directory_list")
            os.chdir(file_path)
            print(f"Directory changed to: {os.getcwd()}")
        else:
            print(f"Yes {current_directory} is set")
        
        # Load the workbook and select the active worksheet
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        # Find the column based on the given heading
        heading_column = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=start_row, column=col).value == heading:
                heading_column = col
                break

        if heading_column is None:
            print(f'Heading "{heading}" not found.')
            return []
        
        # Initialize an empty list to store the employee data labels
        extracted_heading = []
        # Initialize an empty list to store the extracted data
        employee_data = []
        
        # Loop through the specified range of rows
        row = start_row + 1
        
        while row <= stop_row:
            extracted_data = []
            for _ in range(10):
                if row >= stop_row:
                    break
                merged_cell_value = sheet.cell(row=row, column=heading_column).value
                if merged_cell_value != None:
                    emp_info = re.sub(r"\n+", ", ", merged_cell_value, count=0)
                    add_fname_title = "First Name: " + emp_info
                    add_lname_title = re.sub(r",", " ,Last Name:", add_fname_title, count=1)
                    current_emp_info_l = add_lname_title.strip().split(",")
                    for indx in range(len(current_emp_info_l)):
                        label_data = current_emp_info_l[indx].split(":")
                        if label_data[0].strip() not in extracted_heading:
                            extracted_heading.append(label_data[0].strip())
                        if label_data[1] == None:
                            extracted_data.append("None")
                        else: 
                            extracted_data.append(label_data[1].strip())
                    
                    row += 1

            # Advance 1 row before extracting the next set of 10 rows
            if len(extracted_data) > 0:
                # Remove beginning of payperiod date
                extracted_data[5] = re.sub(r"\d\d\d\d-\d\d-\d\d\s-\s","", extracted_data[5])
                
                # Update "Net Pay" to currency format
                v_type = float(extracted_data[(len(extracted_data)-1)].replace('$', ""))
                extracted_data[(len(extracted_data)-1)] = '${:,.2f}'.format(v_type)
                employee_data.append(extracted_data)
            row += 1
    except FileNotFoundError:
        print(f"Error: Directory not found: {file_path}")
        sys.exit(1)
    except NotADirectoryError:
        print(f"Error: Not a directory: {file_path}")
        sys.exit(1)
    except PermissionError:
        print(f"Error: Permission denied to access: {file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)
    return extracted_heading, employee_data


def payroll_taxes(filepath, filename, heading, start_row, stop_row):
    """
    
    This python function extracts data by: 
    
    1. locating a heading and returning the column index 
    2. extracting the values from 6 rows beneath the heading 
       located in the found column 
    3. After extracting 6 rows it advances 4 rows and continues extracting 6 rows 
       until it reach the given end_row to stop 
    
    Extracts data from an Excel file based on a heading and row range.

    Args:
        filepath: The path to the Excel file.
        heading_text: The text of the heading to search for.
        start_row: The row to start searching for the heading.
        end_row: The last row to process in the sheet.

    Returns:
        A list of extracted values, or an empty list if the heading is not found.
        Returns an error message if there is an issue with the excel file.
    """
    # Get the current working directory
    current_directory = os.getcwd()
    print(f"Current directory: {current_directory}")
    
    try:
        
        # Get the current working directory
        directory_list = current_directory.split("\\")
        if filepath not in directory_list:
            print(f"No {filepath} is in directory_list")
            os.chdir(filepath)
            print(f"Directory changed to: {os.getcwd()}")
        else:
            print(f"Yes {current_directory} is set")
        
        # Load the workbook and select the active worksheet
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        # Find the column based on the given heading
        heading_column = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=start_row, column=col).value == heading:
                heading_column = col
                break

        # Column to the Right
        if heading_column is None:
            print(f'Heading "{heading}" not found.')
            return []

        if heading_column > 1:
            right_column_index = heading_column + 1
            # right_column_letter = get_column_letter(right_column_index)
        
        else:
            print(f'Not found.')
            return []                                                                                                                                                       
        # Initialize an empty list to store the extracted data
        payroll_tax = []
        # extracted_payroll_tax = []
        extracted_payroll_tax_label = []

        # Loop through the specified range of rows
        row = start_row + 1
        while row <= stop_row:
            extracted_payroll_tax = []
            for _ in range(6):
                if row > stop_row:
                 break
                v_tax = sheet.cell(row=row, column=right_column_index).value
                v_label = sheet.cell(row=row, column=heading_column).value
                if v_label not in extracted_payroll_tax_label:
                    extracted_payroll_tax_label.append(v_label)
                extracted_payroll_tax.append(f"${v_tax:,.2f}")
                row += 1
            
            # Skip 4 rows
            payroll_tax.append(extracted_payroll_tax)
            row += 4

    except FileNotFoundError:
        print(f"Error: Directory not found: {filepath}")
        sys.exit(1)
    except NotADirectoryError:
        print(f"Error: Not a directory: {filepath}")
        sys.exit(1)
    except PermissionError:
        print(f"Error: Permission denied to access: {filepath}")
        sys.exit(1)
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)
    return extracted_payroll_tax_label, payroll_tax
            
def extract_taxes_paid(filepath, filename, heading):
    """
    Python script using the openpyxl library to accomplish the task of extracting data 
    from 10 merged rows beneath a heading, advancing 1 row, and continuing the extraction 
    until reaching a specified row to stop:
    
    In this script, the extract_data function includes the following parameters:
    
        file_path: The path to your Excel file.
        heading: The heading to identify the column to extract data from.
        start_row: The row where the table heading is located.
        stop_row: The row to stop extraction.
        
    The script first finds the column based on the given heading. Then, it extracts the values 
    from 10 merged rows beneath the heading, advances 1 row, and continues this pattern until it 
    reaches the given stop row.
    """
    # Array for taxes paid
    __taxes = list()
    
    # Get the current working directory
    current_directory = os.getcwd()
    print(f"Current directory: {current_directory}")
    
    try:
        
        # Get the current working directory
        directory_list = current_directory.split("\\")
        if filepath not in directory_list:
            print(f"No {filepath} is in directory_list")
            os.chdir(filepath)
            print(f"Directory changed to: {os.getcwd()}")
        else:
            print(f"Yes {current_directory} is set")
            
        # Load the workbook and select the active sheet
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active

        # Find the row containing the heading
        heading_row = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == heading:
                    heading_row = cell.row
                    break
            if heading_row:
                break

        if heading_row is None:
            print(f"Heading '{heading}' not found.")

        # Print the values from every 10 cells in column J after the heading
        col_j_index = openpyxl.utils.column_index_from_string('J')
        row_index = heading_row + 10

        while True:
            cell_value = sheet.cell(row=row_index, column=col_j_index).value
            if cell_value is None:
                break
            
            if cell_value is not heading:
                taxes = locale.currency(cell_value, grouping=True) 
                __taxes.append(taxes) 
            row_index += 10
            
    except FileNotFoundError:
        print(f"Error: Directory not found: {filepath}")
        sys.exit(1)
    except NotADirectoryError:
        print(f"Error: Not a directory: {filepath}")
        sys.exit(1)
    except PermissionError:
        print(f"Error: Permission denied to access: {filepath}")
        sys.exit(1)
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)
    return __taxes

def combine_list(x, y, z):
    x.extend(y)
    x.append(z)
    return x

def combine_data_lists(x, y, z):
    for i in range(len(x)):
        x[i].extend(y[i])
        x[i].append(z[i])
    return x

def write_file(dirpath, filename, headings, emp_data):
    
    # Create an output file with the name of the original
    filename = re.sub("xlsx", "csv", filename)
    
    # Get the current working directory
    current_directory = os.getcwd()
    print(f"Current directory: {current_directory}")
    
    # Change the current working directory
    try: 
        
        # Get the current working directory
        directory_list = current_directory.split("\\")
        if filepath not in directory_list:
            print(f"No {filepath} is in directory_list")
            os.chdir(filepath)
            print(f"Directory changed to: {os.getcwd()}")
        else:
            print(f"Yes {current_directory} is set")
        
        # Set Pandas dataframe display options
        # Display all rows
        pd.set_option('display.max_rows', None)

        # Display all columns
        pd.set_option('display.max_columns', None)

        # If you also want to adjust the column width
        pd.set_option('display.max_colwidth', None)
    
        # Create the pandas DataFrame 
        df = pd.DataFrame(emp_data, columns = headings) 
        df = df.replace({None: np.nan})  # Replacing None with NaN for missing values
    
        # Drop multiple columns 'A' and 'C'
        drop_columns = ['SSN', 'Hire Date', 'Check Date', 'Department', 'Work Location', 'Home address', 'FED Additional medicare', 'Taxes Paid']
        df.drop(drop_columns, axis=1, inplace=True)
    
        # Move Columns and data to match Plexxis to Zenefits
        column_to_move = df.pop("NetPay")
        df.insert(8, "NetPay", column_to_move)
        column_to_move = df.pop("CA California sdi")
        df.insert(7, "CA California sdi", column_to_move)
        column_to_move = df.pop("FED Medicare")
        df.insert(6, "FED Medicare", column_to_move)
        column_to_move = df.pop("FED Fica")
        df.insert(5, "FED Fica", column_to_move) 
       
        # Update date format to exclude time
        df['Period'] = pd.to_datetime(df['Period']).dt.date
       
        # Update filename replace spaces with underscores
        filtered_filename = filename.replace(" ", "_")
        # print dataframe. 
        df.to_csv(filtered_filename, index=False)
    
    except FileNotFoundError:
        print(f"Error: Directory not found: {dirpath}")
        sys.exit(1)
    except NotADirectoryError:
        print(f"Error: Not a directory: {dirpath}")
        sys.exit(1)
    except PermissionError:
        print(f"Error: Permission denied to access: {dirpath}")
        sys.exit(1)
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)
    return filtered_filename

if __name__ == '__main__':
    
    file_path_input_prompt =  '''
    Please enter the folder were the files\n
    are stored relative to this script 
    (e.g. C:\\<script directory>\\data): 
    '''
    dir_path_input_prompt = '''
    Please enter a <Zenefits>.xlsx files.\n
    Zenefit file name: "
    '''
    user_input_start_row_prompt = '''
    Please enter the row number of the header: "
    '''
    user_input_stop_row_prompt = '''
    Please enter the row number immediately after the last.
    Employee's: "
    '''
    try:
        user_input_filepath = get_filepath_input(file_path_input_prompt)
        print(f"You entered: {user_input_filepath}")
    except ValueError as e:
        print(e)
    
    try:
        user_input_filename = get_filename_input(dir_path_input_prompt)
        print(f"You entered: {dir_path_input_prompt}")
    except ValueError as e:
        print(e)
        
    try:
        user_input_start_row = get_sort_row_index(user_input_start_row_prompt)
    except ValueError as e:
            print(e)
    
    try:
        user_input_stop_row = get_sort_row_index(user_input_stop_row_prompt)
    except ValueError as e:
            print(e)
    
    '''
    Args
    '''
    filepath = user_input_filepath 
    filename = user_input_filename
    start_row = user_input_start_row
    stop_row = user_input_stop_row
    # Extract Employee Headings and Info
    try:
        l_emp_info_heading, l_emp_info = extract_employee_info(filepath, filename, heading='Employees', start_row=7, stop_row=697)
    except ValueError as e:
        print(e)
    # print(l_emp_info_heading)  
   
    # Extract Payroll Headings and Taxes
    try:
        l_payroll_taxes_headings, l_payroll_taxes = payroll_taxes(filepath, filename, heading='Employee-paid Taxes', start_row=7, stop_row=697)
    except ValueError as e:
        print(e)
    # print(len(l_payroll_taxes))
    print(l_payroll_taxes)
    print(l_payroll_taxes_headings)
    
    # Extract Taxes Paid
    try:
        l_taxes_paid = extract_taxes_paid(filepath, filename, heading='Amount')
        v_taxes_paid_heading= 'Taxes Paid'
    except ValueError as e:
        print(e)
    # print(len(l_taxes_paid))
    # print(l_taxes_paid)
    # print(v_taxes_paid_heading)
  

    # Combined Employee info and Tax headings
    try:
        l_combined_headings = combine_list(l_emp_info_heading, l_payroll_taxes_headings, v_taxes_paid_heading)
    except ValueError as e:
        print(e)
    # print(l_combined_headings)

    # Combine Employee info and Payroll and Total taxes paid
    try: 
        l_combined_emp_tax = combine_data_lists(l_emp_info, l_payroll_taxes, l_taxes_paid)
    except ValueError as e:
        print(e)
    # print(l_combined_emp_tax)
    
    # Write *.csv file to directory
    try:
        output_file = write_file(filepath, filename, l_combined_headings, l_combined_emp_tax)
        print(f"CSV file {output_file} has been processed and written to {user_input_filepath}!")
    except ValueError as e:
        print(e)