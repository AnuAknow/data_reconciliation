import openpyxl
import re

def extract_data(file_path, heading, start_row, stop_row):
    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Find the column based on the given heading
    heading_column = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=start_row, column=col).value == heading:
            heading_column = col
            break

    if heading_column is None:
        print(f'Heading "{heading}" not found.')
        return []
    
    # Initialize an empty list to store the employee data labels
    extracted_heading = []
    # Initialize an empty list to store the extracted data
    employee_data = []
    
    # Loop through the specified range of rows
    row = start_row + 1
    
    while row <= stop_row:
        extracted_data = []
        for _ in range(10):
            if row >= stop_row:
                break
            merged_cell_value = sheet.cell(row=row, column=heading_column).value
            if merged_cell_value != None:
                emp_info = re.sub(r"\n+", ", ", merged_cell_value, count=0)
                add_fname_title = "First Name: " + emp_info
                add_lname_title = re.sub(r",", " ,Last Name:", add_fname_title, count=1)
                current_emp_info_l = add_lname_title.strip().split(",")
                for indx in range(len(current_emp_info_l)):
                    label_data = current_emp_info_l[indx].split(":")
                    if label_data[0].strip() not in extracted_heading:
                        extracted_heading.append(label_data[0].strip())
                    if label_data[1] == None:
                        extracted_data.append("None")
                    else: 
                        extracted_data.append(label_data[1].strip())
                
                row += 1

        # Advance 1 row before extracting the next set of 10 rows
        if len(extracted_data) > 0:
            employee_data.append(extracted_data)
        row += 1

    return extracted_heading, employee_data


if __name__ == '__main__':
    v_path = 'data\\'
    v_file = 'Zenefits_payroll_detail_Ck Date 121523.xlsx'
    filepath = v_path+v_file  # Replace with your file path
    heading_text = "Employee-paid Taxes"  # Replace with the actual heading text
    start_row = 7 # Where to start looking for the header
    stop_row = 697 # The last row to process

    headings, data = extract_data(filepath, heading='Employees', start_row=7, stop_row=697)
    print(len(data))
    print(headings)
    for indx in range(len(data)):
        print(data[indx])