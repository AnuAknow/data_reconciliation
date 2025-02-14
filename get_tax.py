import openpyxl
from decimal import Decimal, getcontext

# Set the desired precision (e.g., for typical currency with 2 decimal places)
getcontext().prec = 2
l_taxes = list()
l_taxes_paid = list()
l_emp_info = list()

def print_values_after_heading(file_path, heading):
    __taxes = []
    taxes = ""
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Find the row containing the heading
    heading_row = None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == heading:
                heading_row = cell.row
                break
        if heading_row:
            break
    
    if heading_row is None:
        print(f"Heading '{heading}' not found.")

    # Print the values from every 10 cells in column J after the heading
    col_j_index = openpyxl.utils.column_index_from_string('J')
    col_i_index = openpyxl.utils.column_index_from_string('I')
    row_index = heading_row + 1
    while True:
        cell_value = sheet.cell(row=row_index, column=col_j_index).value
        cell_tax_title = sheet.cell(row=row_index, column=col_i_index).value
        
        if cell_value is None:
            return __taxes 
            
        if cell_tax_title == 'CA California state tax':
            __taxes.append(taxes)
            taxes = ""
            row_index += 4
            
        if cell_value is not None:
            taxes += cell_tax_title + ":" + f"${cell_value:,.2f}" + " ,"
            row_index += 1
            
    
if __name__ == '__main__':
    # Example usage
    l_taxes = print_values_after_heading('data\\Zenefits_payroll_detail_Ck Date 121523.xlsx', 'Amount')
    print(len(l_taxes))
