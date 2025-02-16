from openpyxl import load_workbook
from decimal import getcontext
import locale

# Set the desired precision (e.g., for typical currency with 2 decimal places)
getcontext().prec = 2
locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')

def extract_data_from_excel(filepath, heading, start_row, stop_row):
    """
    Extracts data from an Excel file based on a heading and row range.

    Args:
        filepath: The path to the Excel file.
        heading_text: The text of the heading to search for.
        start_row: The row to start searching for the heading.
        end_row: The last row to process in the sheet.

    Returns:
        A list of extracted values, or an empty list if the heading is not found.
        Returns an error message if there is an issue with the excel file.
    """
    try:
        # Load the workbook and select the active worksheet
        workbook = load_workbook(filepath)
        sheet = workbook.active

        # Find the column based on the given heading
        heading_column = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=start_row, column=col).value == heading:
                heading_column = col
                break

        # Column to the Right
        if heading_column is None:
            print(f'Heading "{heading}" not found.')
            return []

        if heading_column > 1:
            right_column_index = heading_column + 1
            # right_column_letter = get_column_letter(right_column_index)
        
        else:
            print(f'Not found.')
            return []                                                                                                                                                       
        # Initialize an empty list to store the extracted data
        payroll_tax = []
        # extracted_payroll_tax = []
        extracted_payroll_tax_label = []
        
        

        # Loop through the specified range of rows
        row = start_row + 1
        while row <= stop_row:
            extracted_payroll_tax = []
            for _ in range(6):
                if row > stop_row:
                 break
                v_tax = sheet.cell(row=row, column=right_column_index).value
                v_label = sheet.cell(row=row, column=heading_column).value
                if v_label not in extracted_payroll_tax_label:
                    extracted_payroll_tax_label.append(v_label)
                extracted_payroll_tax.append(f"${v_tax:,.2f}")
                row += 1
            
            # Skip 4 rows
            payroll_tax.append(extracted_payroll_tax)
            row += 4

        return extracted_payroll_tax_label, payroll_tax
    except FileNotFoundError:
        return "Error: File not found."
    except Exception as e: # Catch other potential exceptions like invalid excel file
        return f"Error: An error occurred: {e}"


if __name__ == '__main__':
    v_path = 'data\\'
    v_file = 'Zenefits_payroll_detail_Ck Date 121523.xlsx'
    filepath = v_path+v_file  # Replace with your file path
    heading_text = "Employee-paid Taxes"  # Replace with the actual heading text
    start_row = 7 # Where to start looking for the header
    stop_row = 697 # The last row to process
    l_payroll_tax_label, l_payroll_tax = extract_data_from_excel(filepath, heading_text, start_row, stop_row)
    print(len(l_payroll_tax))
    print(l_payroll_tax_label)
    for row in l_payroll_tax:
        print(row)
    # l_extracted_headings, l_extracted_taxes = extract_data_from_excel(filepath, heading_text, start_row, end_row)
    
    # print(l_extracted_taxes)
    # print(l_extracted_headings)
    # if isinstance(l_extracted_taxes, str) and l_extracted_taxes.startswith("Error"):
    #     print(l_extracted_taxes)  # Print the error message
    # else:
    #     if l_extracted_taxes:
    #         print("Extracted values:")
    #         for value in l_extracted_taxes:
    #             print(value)
    #     else:
    #         print("Heading not found or no data extracted.")