from decimal import Decimal, getcontext
from tabulate import tabulate
import pandas as pd 
import numpy as np
import locale
import openpyxl
# import csv
import re


# Set the desired precision (e.g., for typical currency with 2 decimal places)
getcontext().prec = 2

locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
# l_headings = list()
# l_taxes = list()
# l_taxes_paid = list()
# l_emp_info = list()
# d_employee_info = dict()


def extract_employee_info(file_path, heading, start_row, stop_row):
    
    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Find the column based on the given heading
    heading_column = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=start_row, column=col).value == heading:
            heading_column = col
            break

    if heading_column is None:
        print(f'Heading "{heading}" not found.')
        return []
    
    # Initialize an empty list to store the employee data labels
    extracted_heading = []
    # Initialize an empty list to store the extracted data
    employee_data = []
    
    # Loop through the specified range of rows
    row = start_row + 1
    
    while row <= stop_row:
        extracted_data = []
        for _ in range(10):
            if row >= stop_row:
                break
            merged_cell_value = sheet.cell(row=row, column=heading_column).value
            if merged_cell_value != None:
                emp_info = re.sub(r"\n+", ", ", merged_cell_value, count=0)
                add_fname_title = "First Name: " + emp_info
                add_lname_title = re.sub(r",", " ,Last Name:", add_fname_title, count=1)
                current_emp_info_l = add_lname_title.strip().split(",")
                for indx in range(len(current_emp_info_l)):
                    label_data = current_emp_info_l[indx].split(":")
                    if label_data[0].strip() not in extracted_heading:
                        extracted_heading.append(label_data[0].strip())
                    if label_data[1] == None:
                        extracted_data.append("None")
                    else: 
                        extracted_data.append(label_data[1].strip())
                
                row += 1

        # Advance 1 row before extracting the next set of 10 rows
        if len(extracted_data) > 0:
            employee_data.append(extracted_data)
        row += 1

    return extracted_heading, employee_data


def payroll_taxes(filepath, heading, start_row, stop_row):
    """
    Extracts data from an Excel file based on a heading and row range.

    Args:
        filepath: The path to the Excel file.
        heading_text: The text of the heading to search for.
        start_row: The row to start searching for the heading.
        end_row: The last row to process in the sheet.

    Returns:
        A list of extracted values, or an empty list if the heading is not found.
        Returns an error message if there is an issue with the excel file.
    """
    try:
        # Load the workbook and select the active worksheet
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        # Find the column based on the given heading
        heading_column = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=start_row, column=col).value == heading:
                heading_column = col
                break

        # Column to the Right
        if heading_column is None:
            print(f'Heading "{heading}" not found.')
            return []

        if heading_column > 1:
            right_column_index = heading_column + 1
            # right_column_letter = get_column_letter(right_column_index)
        
        else:
            print(f'Not found.')
            return []                                                                                                                                                       
        # Initialize an empty list to store the extracted data
        payroll_tax = []
        # extracted_payroll_tax = []
        extracted_payroll_tax_label = []

        # Loop through the specified range of rows
        row = start_row + 1
        while row <= stop_row:
            extracted_payroll_tax = []
            for _ in range(6):
                if row > stop_row:
                 break
                v_tax = sheet.cell(row=row, column=right_column_index).value
                v_label = sheet.cell(row=row, column=heading_column).value
                if v_label not in extracted_payroll_tax_label:
                    extracted_payroll_tax_label.append(v_label)
                extracted_payroll_tax.append(f"${v_tax:,.2f}")
                row += 1
            
            # Skip 4 rows
            payroll_tax.append(extracted_payroll_tax)
            row += 4

        return extracted_payroll_tax_label, payroll_tax
    except FileNotFoundError:
        return "Error: File not found."
    except Exception as e: # Catch other potential exceptions like invalid excel file
        return f"Error: An error occurred: {e}"
            
            
def extract_taxes_paid(filepath, heading):
    
    # Array for taxes paid
    __taxes = list()
    
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    # Find the row containing the heading
    heading_row = None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == heading:
                heading_row = cell.row
                break
        if heading_row:
            break
    
    if heading_row is None:
        print(f"Heading '{heading}' not found.")

    # Print the values from every 10 cells in column J after the heading
    col_j_index = openpyxl.utils.column_index_from_string('J')
    row_index = heading_row + 10

    
    while True:
        cell_value = sheet.cell(row=row_index, column=col_j_index).value
        if cell_value is None:
            break
        
        if cell_value is not heading:
            taxes = locale.currency(cell_value, grouping=True) 
            __taxes.append(taxes) 
        row_index += 10
        
    return __taxes

def combine_list(x, y, z):
    x.extend(y)
    x.append(z)
    return x

def combine_data_lists(x, y, z):
    for i in range(len(x)):
        x[i].extend(y[i])
        x[i].append(z[i])
    return x

def write_file(dirpath, filename, headings, emp_data):
    
    filename = re.sub("xlsx", "csv", filename)
    filepath = dirpath + filename
    
    # Display all rows
    pd.set_option('display.max_rows', None)

    # Display all columns
    pd.set_option('display.max_columns', None)

    # If you also want to adjust the column width
    pd.set_option('display.max_colwidth', None)
    
    # Create the pandas DataFrame 
    df = pd.DataFrame(emp_data, columns = headings) 
    df = df.replace({None: np.nan})  # Replacing None with NaN for missing values
    
    # print dataframe. 
    df.to_csv(filepath, index=False)

if __name__ == '__main__':
    v_path = 'data\\'
    v_file = 'Zenefits_payroll_detail_Ck Date 121523.xlsx'
    filepath = v_path+v_file  # Replace with your file path
    heading_text = "Employee-paid Taxes"  # Replace with the actual heading text
    start_row = 7 # Where to start looking for the header
    stop_row = 697 # The last row to process

    # Extract Employee Headings and Info
    l_emp_info_heading, l_emp_info = extract_employee_info(filepath, heading='Employees', start_row=7, stop_row=697)
    print(len(l_emp_info))
    print(l_emp_info) 
    print(l_emp_info_heading)  
   
   # Extract Payroll Headings and Taxes
    l_payroll_taxes_headings, l_payroll_taxes = payroll_taxes(filepath, heading='Employee-paid Taxes', start_row=7, stop_row=697)
    print(len(l_payroll_taxes))
    print(l_payroll_taxes)
    print(l_payroll_taxes_headings)
    
    # Extract Taxes Paid
    l_taxes_paid = extract_taxes_paid(filepath, heading='Amount')
    print(len(l_taxes_paid))
    print(l_taxes_paid)
    v_taxes_paid_heading= 'Taxes Paid'
    print(v_taxes_paid_heading)
  

    # Combined Employee info and Tax headings
    # l_combined_headings = combine_list(l_emp_info_heading, l_payroll_taxes_headings, v_taxes_paid_heading)
    # print(l_combined_headings)

    # Combine Employee info and Payroll and Total taxes paid 
    # l_combined_emp_tax = combine_data_lists(l_emp_info, l_payroll_taxes, l_taxes_paid)
    # print(l_combined_emp_tax[1])
    
    # Write *.csv file to directory
    # write_file(v_path, v_file, l_combined_headings, l_combined_emp_tax)
    
    # print(f"CSV file {v_file} has been processed and written to {v_path}!")